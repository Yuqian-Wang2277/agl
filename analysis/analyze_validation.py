#!/usr/bin/env python3
"""Analyze validation outputs and generate Excel report.

Usage:
    python analyze_validation.py <validation_dir> [--output <output.xlsx>]

Examples:
    # Analyze a specific experiment's validation folder
    python analyze_validation.py /path/to/validation_outputs/20260203_221225

    # Specify output Excel path
    python analyze_validation.py /path/to/validation_outputs/20260203_221225 --output results.xlsx

    # The script auto-detects all validation_global_step*.json files in the given directory.
"""

import argparse
import json
import logging
import os
import re
import sys
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ============================================================
# Problem type → category mapping
# ============================================================
ID_TYPES = {
    "contextual_parametric_knowledge_conflicts",
    "cryptonite",
    "disfl_qa",
    "elementary_math_qa",
    "fact_checker",
    "language_identification",
    "matrixshapes",
    "mnist_ascii",
    "movie_dialog_same_or_different",
    "vitaminc_fact_verification",
    "word_unscrambling",
}

OOD_TYPES = {
    "arithmetic",
    "ascii_word_recognition",
    "chess_state_tracking",
    "discourse_marker_prediction",
    "goal_step_wikihow",
    "hyperbaton",
    "implicatures",
    "intersect_geometry",
    "linguistic_mappings",
    "modified_arithmetic",
    "nonsense_words_grammar",
    "real_or_fake_text",
    "simp_turing_concept",
    "snarks",
    "unnatural_in_context_learning",
}

HARD_TYPES = {
    "boolean_expressions",
    "causal_judgement",
    "date_understanding",
    "disambiguation_qa",
    "dyck_languages",
    "formal_fallacies",
    "geometric_shapes",
    "hyperbaton",
    "logical_deduction",
    "movie_recommendation",
    "multistep_arithmetic_two",
    "navigate",
    "object_counting",
    "penguins_in_a_table",
    "reasoning_about_colored_objects",
    "ruin_names",
    "salient_translation_error_detection",
    "snarks",
    "sports_understanding",
    "temporal_sequences",
    "tracking_shuffled_objects",
    "web_of_lies",
    "word_sorting",
}

CATEGORIES = ["ID", "OOD", "HARD"]


# ============================================================
# Data loading
# ============================================================
def load_json_robust(filepath: str) -> List[Dict[str, Any]]:
    """Load a JSON file that may contain a single array (possibly with trailing data)."""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    first_bracket = content.find("[")
    if first_bracket == -1:
        return []

    # Walk through the content respecting strings to find the matching ']'
    stack = 0
    in_string = False
    escape = False
    for i, char in enumerate(content[first_bracket:], first_bracket):
        if escape:
            escape = False
            continue
        if char == "\\":
            escape = True
            continue
        if char == '"' and not escape:
            in_string = not in_string
            continue
        if in_string:
            continue
        if char == "[":
            stack += 1
        elif char == "]":
            stack -= 1
            if stack == 0:
                try:
                    return json.loads(content[first_bracket : i + 1])
                except json.JSONDecodeError as e:
                    logger.error("JSON decode error in %s: %s", filepath, e)
                    return []
    return []


def classify_problem(problem_type: str) -> str:
    """Return 'ID' / 'OOD' / 'HARD' / 'UNKNOWN'."""
    if problem_type in ID_TYPES:
        return "ID"
    if problem_type in OOD_TYPES:
        return "OOD"
    if problem_type in HARD_TYPES:
        return "HARD"
    return "UNKNOWN"


def merge_worker_shards(validation_dir: str) -> None:
    """Find worker shard files without a merged global file and merge them.

    Shard naming: validation_step{N}_worker{PID}.json
    Merged naming: validation_global_step{N}.json

    If merging a step fails (e.g. corrupt JSON), that step is skipped with a warning.
    """
    shard_pattern = re.compile(r"validation_step(\d+)_worker\d+\.json$")
    merged_pattern = re.compile(r"validation_global_step(\d+)\.json$")

    # Collect existing merged steps
    merged_steps: set[int] = set()
    # Collect shard files grouped by step
    shard_map: Dict[int, List[str]] = defaultdict(list)

    for fname in os.listdir(validation_dir):
        m = merged_pattern.match(fname)
        if m:
            merged_steps.add(int(m.group(1)))
            continue
        m = shard_pattern.match(fname)
        if m:
            step = int(m.group(1))
            shard_map[step].append(os.path.join(validation_dir, fname))

    # Merge steps that have shards but no merged file
    for step in sorted(shard_map):
        if step in merged_steps:
            continue
        shards = sorted(shard_map[step])
        merged_data: List[Dict[str, Any]] = []
        ok = True
        for shard_path in shards:
            try:
                items = load_json_robust(shard_path)
                if items:
                    merged_data.extend(items)
            except Exception as e:
                logger.warning("合并 step %d 时读取 %s 失败: %s", step, shard_path, e)
                ok = False
                break
        if not ok or not merged_data:
            print(f"⚠️  Step {step}: 分片合并失败或为空，跳过")
            continue
        out_path = os.path.join(validation_dir, f"validation_global_step{step}.json")
        try:
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(merged_data, f, ensure_ascii=False)
            print(f"🔀 Step {step}: 合并 {len(shards)} 个分片 → {len(merged_data)} 条 ({os.path.basename(out_path)})")
        except Exception as e:
            print(f"⚠️  Step {step}: 写入合并文件失败 ({e})，跳过")


def discover_steps(validation_dir: str) -> List[Tuple[int, str]]:
    """Return sorted list of (step, filepath) found in the directory.

    Automatically merges any unmerged worker shard files before discovery.
    """
    merge_worker_shards(validation_dir)

    pattern = re.compile(r"validation_global_step(\d+)\.json$")
    results = []
    for fname in os.listdir(validation_dir):
        m = pattern.match(fname)
        if m:
            step = int(m.group(1))
            results.append((step, os.path.join(validation_dir, fname)))
    results.sort()
    return results


# ============================================================
# Statistics
# ============================================================
METRIC_KEYS = [
    "final_reward",
    "format_reward",
    "correctness_reward",
    "consistency_reward",
    "coverage_reward",
    "order_reward",
    "binding_reward",
    "intermediate_reward",
]

# Map canonical metric key → list of fallback keys to try in the data
_REWARD_ALIASES: Dict[str, List[str]] = {
    "final_reward": ["final_reward", "final"],
    "format_reward": ["format_reward", "format"],
    "correctness_reward": ["correctness_reward", "correctness"],
    "consistency_reward": ["consistency_reward", "consistency"],
    "coverage_reward": ["coverage_reward", "coverage"],
    "order_reward": ["order_reward", "order"],
    "binding_reward": ["binding_reward", "binding"],
    "intermediate_reward": ["intermediate_reward", "intermediate"],
}


def _get_reward(reward_dict: Dict[str, Any], canonical_key: str, default: float = 0) -> float:
    """Get reward value trying canonical key then alias fallbacks."""
    for alias in _REWARD_ALIASES.get(canonical_key, [canonical_key]):
        if alias in reward_dict:
            return reward_dict[alias]
    return default


def compute_stats(items: List[Dict[str, Any]]) -> Optional[Dict[str, float]]:
    """Compute accuracy / reward metrics for a group of samples."""
    if not items:
        return None
    n = len(items)
    correct = sum(1 for it in items if _get_reward(it["reward"], "correctness_reward") == 1.0)
    fmt_ok = sum(1 for it in items if _get_reward(it["reward"], "format_reward") == 1.0)
    stats: Dict[str, float] = {
        "n": n,
        "acc": correct / n * 100,
        "fmt_acc": fmt_ok / n * 100,
    }
    for key in METRIC_KEYS:
        stats[key] = sum(_get_reward(it["reward"], key) for it in items) / n
    return stats


def analyze_step(data: List[Dict[str, Any]]) -> Dict[str, Optional[Dict[str, float]]]:
    """Group samples by category and compute stats per category."""
    groups: Dict[str, List[Dict[str, Any]]] = {c: [] for c in CATEGORIES}
    for item in data:
        cat = classify_problem(item.get("problem_type", ""))
        if cat in groups:
            groups[cat].append(item)
    return {cat: compute_stats(items) for cat, items in groups.items()}


# ============================================================
# Console output
# ============================================================
def print_report(all_stats: Dict[int, Dict[str, Optional[Dict[str, float]]]], steps: List[int]) -> None:
    """Print human-readable report to stdout."""
    step_labels = "→".join(str(s) for s in steps)

    def trend(cat: str, metric: str, dec: int = 2) -> str:
        vals = []
        for s in steps:
            st = all_stats[s][cat]
            if st is None:
                vals.append("-")
            else:
                vals.append(f"{st[metric]:.{dec}f}")
        return "→".join(vals)

    def delta(cat: str, metric: str) -> str:
        s0, sN = all_stats[steps[0]][cat], all_stats[steps[-1]][cat]
        if s0 and sN:
            return f"{sN[metric] - s0[metric]:+.2f}"
        return "-"

    print("=" * 100)
    print(f"实验结果汇总 (Step: {step_labels})")
    print("=" * 100)

    # Table 1
    print("\n【表1】准确率 (%)")
    print("-" * 80)
    print(f"{'':5} | {'Acc (正确率)':<40} | {'Fmt (格式正确率)':<35}")
    print("-" * 80)
    for cat in CATEGORIES:
        print(f"{cat:<5} | {trend(cat,'acc',1):<35} Δ{delta(cat,'acc'):>6} | {trend(cat,'fmt_acc',1):<30} Δ{delta(cat,'fmt_acc'):>6}")
    print("-" * 80)

    # Table 2
    print("\n【表2】Reward平均值")
    print("-" * 100)
    print(f"{'':5} | {'Final':<37} | {'Fmt':<27} | {'Corr':<27}")
    print("-" * 100)
    for cat in CATEGORIES:
        print(
            f"{cat:<5} | {trend(cat,'final_reward'):<32} Δ{delta(cat,'final_reward'):>6}"
            f" | {trend(cat,'format_reward'):<22} Δ{delta(cat,'format_reward'):>6}"
            f" | {trend(cat,'correctness_reward'):<22} Δ{delta(cat,'correctness_reward'):>6}"
        )
    print("-" * 100)
    print(f"\n{'':5} | {'Cons (一致性)':<40}")
    print("-" * 55)
    for cat in CATEGORIES:
        print(f"{cat:<5} | {trend(cat,'consistency_reward'):<35} Δ{delta(cat,'consistency_reward'):>6}")
    print("-" * 55)

    # Table 3
    print("\n【表3】Consistency子指标 (平均值)")
    print("-" * 105)
    print(f"{'':5} | {'Cov':<24} | {'Ord':<24} | {'Bind':<24} | {'Inter':<22}")
    print("-" * 105)
    for cat in CATEGORIES:
        print(
            f"{cat:<5} | {trend(cat,'coverage_reward'):<24}"
            f" | {trend(cat,'order_reward'):<24}"
            f" | {trend(cat,'binding_reward'):<24}"
            f" | {trend(cat,'intermediate_reward'):<22}"
        )
    print("-" * 105)

    # Table 4
    print("\n【表4】验证样本数 (每step实际采样数)")
    print("-" * (14 + 9 * len(steps)))
    header = f"{'Step':<8}" + "".join(f" | {s:>5}" for s in steps)
    print(header)
    print("-" * (14 + 9 * len(steps)))
    for cat in CATEGORIES:
        vals = [all_stats[s][cat]["n"] if all_stats[s][cat] else 0 for s in steps]
        row = f"{cat:<8}" + "".join(f" | {int(v):>5}" for v in vals)
        print(row)
    totals = [sum(all_stats[s][c]["n"] for c in CATEGORIES if all_stats[s][c]) for s in steps]
    print("-" * (14 + 9 * len(steps)))
    total_row = f"{'Total':<8}" + "".join(f" | {int(t):>5}" for t in totals)
    print(total_row)
    print("-" * (14 + 9 * len(steps)))


# ============================================================
# Excel output
# ============================================================
# Style constants
HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
CAT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row: int, col_start: int, col_end: int) -> None:
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_cat(ws, row: int, col: int) -> None:
    cell = ws.cell(row=row, column=col)
    cell.font = Font(bold=True, size=10)
    cell.fill = CAT_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def _write_trend_cell(ws, row: int, col: int, values: List[Optional[float]], fmt: str = ".1f") -> None:
    """Write trend value as string in a single cell."""
    parts = [f"{v:{fmt}}" if v is not None else "-" for v in values]
    cell = ws.cell(row=row, column=col, value="→".join(parts))
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def _write_delta_cell(ws, row: int, col: int, v0: Optional[float], vN: Optional[float], fmt: str = ".2f") -> None:
    if v0 is not None and vN is not None:
        d = vN - v0
        cell = ws.cell(row=row, column=col, value=f"{d:+{fmt}}")
    else:
        cell = ws.cell(row=row, column=col, value="-")
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def _get_vals(all_stats: Dict, steps: List[int], cat: str, metric: str) -> List[Optional[float]]:
    return [all_stats[s][cat][metric] if all_stats[s][cat] else None for s in steps]


def generate_excel(
    all_stats: Dict[int, Dict[str, Optional[Dict[str, float]]]],
    steps: List[int],
    output_path: str,
    all_data: Optional[Dict[int, List[Dict[str, Any]]]] = None,
) -> None:
    """Generate Excel workbook with all analysis tables."""
    wb = Workbook()
    step_label = "→".join(str(s) for s in steps)

    # ---- Sheet 1: 准确率 ----
    ws1 = wb.active
    ws1.title = "准确率"

    # Row 1: top-level header
    ws1.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws1.cell(row=1, column=1, value="Cat")
    ws1.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
    ws1.cell(row=1, column=2, value="准确率指标 (%)")
    _style_header(ws1, 1, 1, 3)

    # Row 2: sub headers
    ws1.cell(row=2, column=2, value=f"Acc ({step_label})")
    ws1.cell(row=2, column=3, value=f"Fmt ({step_label})")
    _style_header(ws1, 2, 1, 3)

    for i, cat in enumerate(CATEGORIES):
        r = 3 + i
        _style_cat(ws1, r, 1)
        ws1.cell(row=r, column=1, value=cat)
        _write_trend_cell(ws1, r, 2, _get_vals(all_stats, steps, cat, "acc"), ".1f")
        _write_trend_cell(ws1, r, 3, _get_vals(all_stats, steps, cat, "fmt_acc"), ".1f")

    # Δ row
    ws1.cell(row=6, column=1, value="Δ (末-初)")
    _style_header(ws1, 6, 1, 3)
    for i, cat in enumerate(CATEGORIES):
        r = 7 + i
        _style_cat(ws1, r, 1)
        ws1.cell(row=r, column=1, value=cat)
        _write_delta_cell(ws1, r, 2, _get_vals(all_stats, steps, cat, "acc")[0], _get_vals(all_stats, steps, cat, "acc")[-1])
        _write_delta_cell(ws1, r, 3, _get_vals(all_stats, steps, cat, "fmt_acc")[0], _get_vals(all_stats, steps, cat, "fmt_acc")[-1])

    for col in range(1, 4):
        ws1.column_dimensions[get_column_letter(col)].width = 40

    # ---- Sheet 2: Reward指标 ----
    ws2 = wb.create_sheet("Reward指标")

    reward_metrics = [
        ("Final", "final_reward"),
        ("Fmt", "format_reward"),
        ("Corr", "correctness_reward"),
        ("Cons", "consistency_reward"),
    ]

    ws2.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws2.cell(row=1, column=1, value="Cat")
    ws2.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(reward_metrics) + 1)
    ws2.cell(row=1, column=2, value="Reward平均值")
    _style_header(ws2, 1, 1, len(reward_metrics) + 1)

    for j, (label, _) in enumerate(reward_metrics):
        ws2.cell(row=2, column=2 + j, value=f"{label} ({step_label})")
    _style_header(ws2, 2, 1, len(reward_metrics) + 1)

    for i, cat in enumerate(CATEGORIES):
        r = 3 + i
        _style_cat(ws2, r, 1)
        ws2.cell(row=r, column=1, value=cat)
        for j, (_, key) in enumerate(reward_metrics):
            _write_trend_cell(ws2, r, 2 + j, _get_vals(all_stats, steps, cat, key), ".2f")

    # Δ row
    ws2.cell(row=6, column=1, value="Δ (末-初)")
    _style_header(ws2, 6, 1, len(reward_metrics) + 1)
    for i, cat in enumerate(CATEGORIES):
        r = 7 + i
        _style_cat(ws2, r, 1)
        ws2.cell(row=r, column=1, value=cat)
        for j, (_, key) in enumerate(reward_metrics):
            vals = _get_vals(all_stats, steps, cat, key)
            _write_delta_cell(ws2, r, 2 + j, vals[0], vals[-1])

    for col in range(1, len(reward_metrics) + 2):
        ws2.column_dimensions[get_column_letter(col)].width = 38

    # ---- Sheet 3: Consistency子指标 ----
    ws3 = wb.create_sheet("Consistency子指标")

    sub_metrics = [
        ("Cov", "coverage_reward"),
        ("Ord", "order_reward"),
        ("Bind", "binding_reward"),
        ("Inter", "intermediate_reward"),
    ]

    ws3.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws3.cell(row=1, column=1, value="Cat")
    ws3.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(sub_metrics) + 1)
    ws3.cell(row=1, column=2, value="Consistency子指标 (平均值)")
    _style_header(ws3, 1, 1, len(sub_metrics) + 1)

    for j, (label, _) in enumerate(sub_metrics):
        ws3.cell(row=2, column=2 + j, value=f"{label} ({step_label})")
    _style_header(ws3, 2, 1, len(sub_metrics) + 1)

    for i, cat in enumerate(CATEGORIES):
        r = 3 + i
        _style_cat(ws3, r, 1)
        ws3.cell(row=r, column=1, value=cat)
        for j, (_, key) in enumerate(sub_metrics):
            _write_trend_cell(ws3, r, 2 + j, _get_vals(all_stats, steps, cat, key), ".2f")

    for col in range(1, len(sub_metrics) + 2):
        ws3.column_dimensions[get_column_letter(col)].width = 38

    # ---- Sheet 4: 样本数 ----
    ws4 = wb.create_sheet("样本数")

    ws4.cell(row=1, column=1, value="Cat \\ Step")
    for j, s in enumerate(steps):
        ws4.cell(row=1, column=2 + j, value=f"Step {s}")
    _style_header(ws4, 1, 1, 1 + len(steps))

    for i, cat in enumerate(CATEGORIES):
        r = 2 + i
        _style_cat(ws4, r, 1)
        ws4.cell(row=r, column=1, value=cat)
        for j, s in enumerate(steps):
            st = all_stats[s][cat]
            cell = ws4.cell(row=r, column=2 + j, value=int(st["n"]) if st else 0)
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    # Total row
    r_total = 2 + len(CATEGORIES)
    ws4.cell(row=r_total, column=1, value="Total")
    _style_header(ws4, r_total, 1, 1 + len(steps))
    for j, s in enumerate(steps):
        total = sum(all_stats[s][c]["n"] for c in CATEGORIES if all_stats[s][c])
        cell = ws4.cell(row=r_total, column=2 + j, value=int(total))
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws4.column_dimensions["A"].width = 12
    for j in range(len(steps)):
        ws4.column_dimensions[get_column_letter(2 + j)].width = 10

    # ---- Sheet 5: 题型细分 ----
    if all_data:
        LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
        last_step = steps[-1]
        data = all_data.get(last_step, [])
        if data:
            ws5 = wb.create_sheet("题型细分")
            ws5.cell(row=1, column=1, value=f"Step {last_step} 各题型细分")
            ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
            _style_header(ws5, 1, 1, 7)

            headers = ["Cat", "Problem Type", "N", "Acc (%)", "Avg Final", "Avg Fmt", "Avg Corr"]
            for j, h in enumerate(headers):
                ws5.cell(row=2, column=1 + j, value=h)
            _style_header(ws5, 2, 1, 7)

            # Group by category then problem_type
            cat_type_items: Dict[str, Dict[str, List[Dict[str, Any]]]] = {c: defaultdict(list) for c in CATEGORIES}
            for item in data:
                cat = classify_problem(item.get("problem_type", ""))
                if cat in cat_type_items:
                    cat_type_items[cat][item.get("problem_type", "unknown")].append(item)

            r = 3
            for cat in CATEGORIES:
                type_groups = cat_type_items[cat]
                if not type_groups:
                    _style_cat(ws5, r, 1)
                    ws5.cell(row=r, column=1, value=cat)
                    ws5.cell(row=r, column=2, value="(no data)")
                    r += 1
                    continue
                cat_n = 0
                cat_correct = 0
                cat_final = 0.0
                cat_fmt = 0.0
                cat_corr = 0.0
                for ptype in sorted(type_groups):
                    items = type_groups[ptype]
                    tn = len(items)
                    tc = sum(1 for it in items if _get_reward(it.get("reward", {}), "correctness_reward") == 1.0)
                    t_final = sum(_get_reward(it.get("reward", {}), "final_reward") for it in items) / tn
                    t_fmt = sum(_get_reward(it.get("reward", {}), "format_reward") for it in items) / tn
                    t_corr = sum(_get_reward(it.get("reward", {}), "correctness_reward") for it in items) / tn

                    _style_cat(ws5, r, 1)
                    ws5.cell(row=r, column=1, value=cat)
                    c2 = ws5.cell(row=r, column=2, value=ptype)
                    c2.alignment = LEFT_ALIGN
                    c2.border = THIN_BORDER
                    for ci, val in [(3, tn), (4, round(tc / tn * 100, 1)), (5, round(t_final, 4)),
                                    (6, round(t_fmt, 4)), (7, round(t_corr, 4))]:
                        c = ws5.cell(row=r, column=ci, value=val)
                        c.alignment = CENTER
                        c.border = THIN_BORDER

                    cat_n += tn
                    cat_correct += tc
                    cat_final += t_final * tn
                    cat_fmt += t_fmt * tn
                    cat_corr += t_corr * tn
                    r += 1

                # Category subtotal
                _style_header(ws5, r, 1, 7)
                ws5.cell(row=r, column=1, value=cat)
                ws5.cell(row=r, column=2, value="小计")
                ws5.cell(row=r, column=3, value=cat_n)
                ws5.cell(row=r, column=4, value=round(cat_correct / cat_n * 100, 1) if cat_n else 0)
                ws5.cell(row=r, column=5, value=round(cat_final / cat_n, 4) if cat_n else 0)
                ws5.cell(row=r, column=6, value=round(cat_fmt / cat_n, 4) if cat_n else 0)
                ws5.cell(row=r, column=7, value=round(cat_corr / cat_n, 4) if cat_n else 0)
                r += 1

            ws5.column_dimensions["A"].width = 8
            ws5.column_dimensions["B"].width = 48
            for ci in range(3, 8):
                ws5.column_dimensions[get_column_letter(ci)].width = 14

    # Save
    wb.save(output_path)
    print(f"\n✅ Excel文件已保存: {output_path}")


# ============================================================
# Main
# ============================================================
def main() -> None:
    parser = argparse.ArgumentParser(
        description="分析验证结果并生成Excel报告",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("validation_dir", help="验证输出文件夹路径 (包含validation_global_step*.json)")
    parser.add_argument("--output", "-o", default=None, help="输出Excel文件路径 (默认: 同目录下 analysis_report.xlsx)")
    args = parser.parse_args()

    validation_dir = os.path.abspath(args.validation_dir)
    if not os.path.isdir(validation_dir):
        print(f"❌ 目录不存在: {validation_dir}", file=sys.stderr)
        sys.exit(1)

    # Discover step files
    step_files = discover_steps(validation_dir)
    if not step_files:
        print(f"❌ 未找到 validation_global_step*.json 文件: {validation_dir}", file=sys.stderr)
        sys.exit(1)

    # Auto-select up to 6 equidistant points if too many steps
    MAX_POINTS = 6
    if len(step_files) > MAX_POINTS:
        n = len(step_files)
        indices = [round(i * (n - 1) / (MAX_POINTS - 1)) for i in range(MAX_POINTS)]
        seen: set[int] = set()
        unique_indices: list[int] = []
        for idx in indices:
            if idx not in seen:
                seen.add(idx)
                unique_indices.append(idx)
        step_files = [step_files[i] for i in unique_indices]

    steps = [s for s, _ in step_files]
    print(f"📂 验证目录: {validation_dir}")
    print(f"📊 分析 {len(steps)} 个step点位: {steps}")

    # Load & analyze
    all_stats: Dict[int, Dict[str, Optional[Dict[str, float]]]] = {}
    all_data: Dict[int, List[Dict[str, Any]]] = {}
    for step, filepath in step_files:
        data = load_json_robust(filepath)
        if not data:
            print(f"⚠️  Step {step}: 加载失败或为空")
            all_stats[step] = {c: None for c in CATEGORIES}
            all_data[step] = []
            continue
        all_data[step] = data
        all_stats[step] = analyze_step(data)
        total = sum(all_stats[step][c]["n"] for c in CATEGORIES if all_stats[step][c])
        print(f"   Step {step}: {total} 样本 (ID={all_stats[step]['ID']['n'] if all_stats[step]['ID'] else 0}, "
              f"OOD={all_stats[step]['OOD']['n'] if all_stats[step]['OOD'] else 0}, "
              f"HARD={all_stats[step]['HARD']['n'] if all_stats[step]['HARD'] else 0})")

    # Print console report
    print_report(all_stats, steps)

    # Generate Excel
    if args.output:
        output_path = os.path.abspath(args.output)
    else:
        output_path = os.path.join(validation_dir, "analysis_report.xlsx")
    generate_excel(all_stats, steps, output_path, all_data)


if __name__ == "__main__":
    main()


